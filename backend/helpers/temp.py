import requests
import openpyxl
from io import BytesIO

def parse_camera_locations_excel(url):
    """
    Downloads the Excel file from the specified URL, parses it using openpyxl,
    and prints out rows from the first worksheet.
    """
    # Step 1: Download the file
    response = requests.get(url)
    response.raise_for_status()  # Raise an exception if the request failed

    # Step 2: Load the workbook from memory
    excel_data = BytesIO(response.content)
    workbook = openpyxl.load_workbook(excel_data, data_only=True)

    # Step 3: Choose the first worksheet (or pick by name)
    sheet = workbook.worksheets[0]
    
    # Optionally print out the sheet name:
    print(f"Parsing data from sheet: {sheet.title}")

    # Step 4: Iterate through rows
    # iter_rows with values_only=True returns tuples with cell values only
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Print out each row; you can also process/store as needed
        print(row_number, row)

if __name__ == "__main__":
    excel_url = "https://www.vic.gov.au/sites/default/files/2024-12/DDS_camera_locations_January-2025.xlsx"
    parse_camera_locations_excel(excel_url)
